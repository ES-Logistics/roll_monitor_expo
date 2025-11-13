"""
Formatador Excel - Responsável pela estruturação e formatação visual dos relatórios
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import json


class ExcelFormatter:
    """Classe responsável pela formatação visual do relatório Excel"""
    
    # Definição das cores (seguindo o exemplo da imagem)
    CURRENT_STATE_COLOR = "ADD8E6"  # Azul claro
    HISTORY_COLOR = "FDBCB4"        # Tom de pele
    
    # Definição das colunas do relatório
    COLUMNS = [
        'Processo', 'Tipo Alteração', 'Cliente', 'Armador', 'Porto Embarque', 'Navio Embarque', 'Embarque',
        'Porto Destino', 'Embarque Transbordo', 'Porto Transbordo', 'Navio Transbordo', 'Email Responsável',
        'Motivo Transferência', 'Versão', 'Quantidade de Alterações'
    ]
    
    def __init__(self):
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Relatorio Alteracoes"
        self.current_row = 1
        self.groups_created = []  # Lista para rastrear grupos criados
    
    def create_report(self, report_data):
        """Cria o relatório Excel completo com agrupamentos"""
        self._setup_header()
        self._add_data_rows(report_data)
        self._apply_groupings()  # Nova função para aplicar agrupamentos
        self._adjust_column_widths()
        return self.workbook
    
    def _setup_header(self):
        """Configura o cabeçalho do relatório"""
        for col_num, column_name in enumerate(self.COLUMNS, 1):
            cell = self.worksheet.cell(row=1, column=col_num)
            cell.value = column_name
            cell.font = Font(bold=True, size=12)
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True, size=12)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self._get_border()
        
        self.current_row = 2
    
    def _add_data_rows(self, report_data):
        """Adiciona todas as linhas de dados ao relatório"""
        for process_data in report_data:
            self._add_process_rows(process_data)
    
    def _add_process_rows(self, process_data):
        """Adiciona as linhas de um processo específico (atual + histórico)"""
        proceso = process_data['proceso']
        current_state = process_data['current_state']
        alteracoes = process_data['alteracoes']
        
        # Extrai histórico de alterações e ordena do mais recente para o mais antigo
        history = alteracoes.get('historico_alteracoes', [])
        if not history:
            return
        
        # Ordena por timestamp (mais recente primeiro)
        history_sorted = sorted(history, key=lambda x: x.get('timestamp', ''), reverse=True)
        total_changes = len(history_sorted)
        
        # NOVA REGRA: Sempre mostra linha azul + versão 0 "Original"
        # Se múltiplas mudanças, também mostra histórico intermediário
        
        # Sempre adiciona linha atual (azul claro) - versão mais alta
        current_version = total_changes
        self._add_current_state_row(proceso, current_state, current_version, total_changes, alteracoes)
        
        # Marca início do grupo de linhas bege para este processo
        group_start_row = self.current_row
        
        # Adiciona histórico intermediário (tom de pele) - do mais recente para o mais antigo
        # Mas pula o mais antigo, pois ele será usado para construir a versão original
        if total_changes > 1:
            for i, change in enumerate(history_sorted[:-1]):  # [:-1] pula o último (mais antigo)
                version = total_changes - i - 1
                self._add_history_row("", change, version, total_changes)
        
        # SEMPRE adiciona versão 0 "Original" baseada no estado ANTES da primeira mudança
        oldest_change = history_sorted[-1]  # Mudança mais antiga
        self._add_original_version_row("", oldest_change, total_changes)
        
        # Marca fim do grupo e registra para agrupamento posterior
        group_end_row = self.current_row - 1
        
        # Só cria grupo se há pelo menos 1 linha bege (sempre há pelo menos a Original)
        if group_end_row >= group_start_row:
            self.groups_created.append({
                'start': group_start_row,
                'end': group_end_row,
                'proceso': proceso
            })
    
    def _add_current_state_row(self, proceso, current_state, version, total_changes, alteracoes_info=None):
        """Adiciona linha com estado atual (azul claro)"""
        if not current_state:
            return
        
        # Determina tipo de alteração e campos alterados baseado no histórico
        tipo_alteracao = ""
        all_changed_fields = set()
        if alteracoes_info and 'historico_alteracoes' in alteracoes_info:
            # Pega todas as alterações e lista os campos únicos alterados
            for change in alteracoes_info['historico_alteracoes']:
                if 'alteracoes' in change:
                    all_changed_fields.update(change['alteracoes'].keys())
            tipo_alteracao = ", ".join(sorted(all_changed_fields))
        
        row_data = [
            proceso,
            tipo_alteracao,  # Mostra campos alterados na linha azul
            current_state.get('cliente', ''),
            current_state.get('armador', ''),
            current_state.get('porto_embarque', ''),
            current_state.get('navio_embarque', ''),
            self._format_datetime(current_state.get('previsao_embarque')),
            current_state.get('porto_destino', ''),
            self._format_datetime(current_state.get('previsao_embarque_transbordo')),
            current_state.get('porto_transbordo', ''),
            current_state.get('navio_transbordo', ''),
            current_state.get('email_responsavel', ''),
            current_state.get('motivo_transferencia', ''),
            f".{version}" if version > 0 else "",
            total_changes
        ]
        
        # NOVA LÓGICA: Passa campos alterados para destacar na linha azul também
        self._write_row(row_data, self.CURRENT_STATE_COLOR, is_current=True, changed_fields=list(all_changed_fields))
    
    def _add_history_row(self, proceso, change, version, total_changes):
        """Adiciona linha de histórico (tom de pele)"""
        # NOVA LÓGICA: Extrai dados da estrutura real
        alteracoes_data = change.get('alteracoes', {})
        
        # Determina quais campos foram alterados
        changed_fields = list(alteracoes_data.keys())
        tipo_alteracao = ", ".join(changed_fields) if changed_fields else ""
        
        # Extrai valores anteriores da estrutura real
        previous_values = {}
        for campo, mudancas_list in alteracoes_data.items():
            if mudancas_list and len(mudancas_list) > 0:
                # Pega o valor anterior da primeira mudança
                previous_values[campo] = mudancas_list[0].get('anterior', '')
        
        row_data = [
            proceso,  # Vazio para histórico (processo só na linha azul)
            tipo_alteracao,
            previous_values.get('cliente', ''),
            previous_values.get('armador', ''),
            previous_values.get('porto_embarque', ''),
            previous_values.get('navio_embarque', ''),
            self._format_datetime(previous_values.get('previsao_embarque')),
            previous_values.get('porto_destino', ''),
            self._format_datetime(previous_values.get('previsao_embarque_transbordo')),
            previous_values.get('porto_transbordo', ''),
            previous_values.get('navio_transbordo', ''),
            previous_values.get('email_responsavel', ''),
            previous_values.get('motivo_transferencia', ''),
            f".{version}",
            total_changes
        ]
        
        self._write_row(row_data, self.HISTORY_COLOR, is_current=False, changed_fields=changed_fields)
    
    def _add_original_version_row(self, proceso, oldest_change, total_changes):
        """Adiciona linha da versão 0 'Original' (tom de pele)"""
        # NOVA LÓGICA: Mostra o estado ANTES da primeira mudança
        alteracoes_data = oldest_change.get('alteracoes', {})
        
        # Para a versão original, pegamos os valores ANTERIORES da primeira mudança
        original_values = {}
        for campo, mudancas_list in alteracoes_data.items():
            if mudancas_list and len(mudancas_list) > 0:
                # O valor "anterior" da primeira mudança é o estado original
                original_values[campo] = mudancas_list[0].get('anterior', '')
        
        row_data = [
            proceso,  # Vazio para histórico
            "Original",  # Tipo alteração sempre "Original"
            original_values.get('cliente', ''),
            original_values.get('armador', ''),
            original_values.get('porto_embarque', ''),
            original_values.get('navio_embarque', ''),
            self._format_datetime(original_values.get('previsao_embarque')),
            original_values.get('porto_destino', ''),
            self._format_datetime(original_values.get('previsao_embarque_transbordo')),
            original_values.get('porto_transbordo', ''),
            original_values.get('navio_transbordo', ''),
            original_values.get('email_responsavel', ''),
            original_values.get('motivo_transferencia', ''),
            ".0",  # Sempre versão 0
            total_changes
        ]
        
        # Não aplica formatação especial na versão original (é referência base)
        self._write_row(row_data, self.HISTORY_COLOR, is_current=False, changed_fields=None)
    
    def _apply_groupings(self):
        """Aplica agrupamentos de linhas e deixa colapsados inicialmente"""
        print(f"🔗 Aplicando {len(self.groups_created)} agrupamentos...")
        
        for group in self.groups_created:
            start_row = group['start']
            end_row = group['end']
            proceso = group['proceso']
            
            try:
                # Cria agrupamento de linhas no Excel
                self.worksheet.row_dimensions.group(start_row, end_row, outline_level=1, hidden=True)
                
                print(f"   📁 Processo {proceso}: linhas {start_row}-{end_row} agrupadas (colapsado)")
                
            except Exception as e:
                print(f"   ⚠️  Erro ao agrupar {proceso}: {e}")
        
        # Configuração adicional do outline
        if self.groups_created:
            # Define que os grupos devem aparecer colapsados por padrão
            self.worksheet.sheet_properties.outlinePr.summaryBelow = True
            self.worksheet.sheet_properties.outlinePr.summaryRight = False
    
    def _write_row(self, row_data, background_color, is_current=False, changed_fields=None):
        """Escreve uma linha no Excel com formatação adequada"""
        for col_num, value in enumerate(row_data, 1):
            cell = self.worksheet.cell(row=self.current_row, column=col_num)
            cell.value = value
            cell.fill = PatternFill(start_color=background_color, end_color=background_color, fill_type="solid")
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = self._get_border()
            
            # NOVA LÓGICA: Aplica formatação especial para campos alterados em AMBAS as linhas (azul e bege)
            if changed_fields and col_num >= 3:  # Pula Processo e Tipo Alteração
                column_name = self._get_field_name_by_column(col_num)
                if column_name and column_name in changed_fields:
                    # FORMATAÇÃO ESPECIAL: Negrito + Itálico + Sublinhado
                    cell.font = Font(bold=True, italic=True, underline="single", color="000000")
                else:
                    cell.font = Font(color="000000")
            else:
                cell.font = Font(color="000000")
        
        self.current_row += 1
    
    def _get_field_name_by_column(self, col_num):
        """Mapeia número da coluna para nome do campo na base"""
        # NOTA: cliente, armador e email_responsavel NÃO triggam eventos de mudança
        # então não são incluídos no mapeamento para formatação
        mapping = {
            5: 'porto_embarque',        # Coluna Porto Embarque
            6: 'navio_embarque',        # Coluna Navio Embarque
            7: 'previsao_embarque',     # Coluna Embarque
            8: 'porto_destino',         # Coluna Porto Destino
            9: 'previsao_embarque_transbordo',  # Coluna Embarque Transbordo
            10: 'porto_transbordo',     # Coluna Porto Transbordo
            11: 'navio_transbordo',     # Coluna Navio Transbordo
            # email_responsavel (coluna 12) não trigga mudanças
            13: 'motivo_transferencia'  # Coluna Motivo Transferência
        }
        return mapping.get(col_num, '')
    
    def _format_datetime(self, datetime_value):
        """Formata valores de data/hora para exibição (APENAS DATA)"""
        if not datetime_value or datetime_value == '' or datetime_value is None:
            return ""
        
        if isinstance(datetime_value, str):
            try:
                # Remove timezone info se presente
                datetime_str = datetime_value.replace('Z', '').replace('+00:00', '')
                # Tenta parsear diferentes formatos
                from datetime import datetime as dt
                try:
                    # Formato ISO completo
                    parsed_dt = dt.fromisoformat(datetime_str)
                except:
                    try:
                        # Formato só com data
                        parsed_dt = dt.strptime(datetime_str[:10], '%Y-%m-%d')
                    except:
                        # Se não conseguir parsear, retorna a string original truncada
                        return str(datetime_value)[:10] if len(str(datetime_value)) >= 10 else str(datetime_value)
                
                return parsed_dt.strftime('%Y-%m-%d')
                
            except Exception:
                # Em caso de erro, retorna string segura
                return str(datetime_value)[:10] if len(str(datetime_value)) >= 10 else str(datetime_value)
                
        elif hasattr(datetime_value, 'strftime'):
            try:
                return datetime_value.strftime('%Y-%m-%d')
            except:
                return str(datetime_value)
        else:
            return str(datetime_value) if datetime_value else ""
    
    def _get_border(self):
        """Define bordas padrão para células"""
        thin_border = Side(border_style="thin", color="000000")
        return Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
    
    def _adjust_column_widths(self):
        """Ajusta largura das colunas automaticamente"""
        for col_num in range(1, len(self.COLUMNS) + 1):
            column_letter = get_column_letter(col_num)
            
            # Define larguras específicas baseadas no conteúdo
            if col_num == 1:  # Processo
                self.worksheet.column_dimensions[column_letter].width = 15
            elif col_num == 2:  # Tipo Alteração
                self.worksheet.column_dimensions[column_letter].width = 20
            elif col_num in [3, 4]:  # Cliente, Armador
                self.worksheet.column_dimensions[column_letter].width = 25
            elif col_num in [7, 9]:  # Datas (Embarque, Embarque Transbordo)
                self.worksheet.column_dimensions[column_letter].width = 15
            elif col_num == 12:  # Email Responsável
                self.worksheet.column_dimensions[column_letter].width = 30
            elif col_num == 13:  # Motivo Transferência
                self.worksheet.column_dimensions[column_letter].width = 25
            elif col_num in [14, 15]:  # Versão e Quantidade
                self.worksheet.column_dimensions[column_letter].width = 10
            else:  # Porto, Navio, etc.
                self.worksheet.column_dimensions[column_letter].width = 18