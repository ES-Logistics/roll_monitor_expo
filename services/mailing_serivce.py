import time
import config
import logging
import base64
import pandas as pd
import io
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from services.query_service import QueryService
from controllers.mail_controller import EmailManager


class MailReport:
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        logging.basicConfig(level=logging.INFO)
        self.recipient_email = config.TARGET_MAILS
        self.subject = "Relatório Roll Monitor Exportação " + time.strftime("%Y-%m-%d %H:%M:%S")
        self.queries = QueryService()
        self.mailing = EmailManager()

    def expand_row(self, row):
        """
        Cria todas as colunas sempre, preenchendo com None quando não houver dados
        """
        # Campos base sempre presentes
        data = {
            "processo": row["processo"],
            "cliente": row["cliente"],
            "armador": row["armador"],
            "campo": row["campo"],
            "email_responsavel": row["email_responsavel"],
            "email_resp_booking": row["email_resp_booking"],
            "email_cs": row["email_cs"],
            "motivo_transferencia": row["motivo_transferencia"],
            "booking": row["booking"],
            "ds_quantidade_containers": row["ds_quantidade_containers"],
            "porto_embarque": row["porto_embarque"]
        }
        
        # Campos específicos - sempre criar as colunas, independente do conteúdo
        campo = row["campo"]
        
        # Lista de todos os campos possíveis que podem ter mudança
        all_possible_fields = [
            "navio_embarque",
            "viagem", 
            "previsao_embarque",
            "navio_transbordo",
            "previsao_embarque_transbordo"
        ]
        
        # Criar colunas (old) e (new) para TODOS os campos, preenchendo com None
        for field in all_possible_fields:
            data[f"{field} (old)"] = None
            data[f"{field} (new)"] = None
        
        # Preencher apenas o campo que realmente mudou
        if campo in all_possible_fields:
            data[f"{campo} (old)"] = row["old_value"]
            data[f"{campo} (new)"] = row["new_value"]
        
        return data
    
    def apply_row_colors_by_qtd_alteracoes(self, ws, qtd_col_name="qtd_alteracoes"):
        """
        Aplica cor de fundo na linha inteira baseado na coluna qtd_alteracoes
        >= 4  -> vermelho pastel
        <= 3  -> azul pastel
        """
        headers = [cell.value for cell in ws[1]]

        if qtd_col_name not in headers:
            return  # nada pra fazer, segue a vida

        qtd_col_idx = headers.index(qtd_col_name) + 1

        red_fill = PatternFill(
            start_color="FDE2E2",
            end_color="FDE2E2",
            fill_type="solid"
        )

        blue_fill = PatternFill(
            start_color="E8F1FA",
            end_color="E8F1FA",
            fill_type="solid"
        )
        yellow_fill = PatternFill(
            start_color="FFF9E5",
            end_color="FFF9E5",
            fill_type="solid"
        )

        for row in range(2, ws.max_row + 1):
            qtd = ws.cell(row=row, column=qtd_col_idx).value

            if qtd is None:
                continue

            fill = red_fill if qtd >= 4 else yellow_fill if qtd > 2 else blue_fill

            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = fill

    def df_to_excel_bytes(self,df):
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Relatorio")
        output.seek(0)
        return output.read()

    def auto_adjust_column_width(self, ws, min_width=10, max_width=50):
        """
        Ajusta automaticamente a largura das colunas com base no conteúdo
        (somente horizontal)
        """
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)

            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            adjusted_width = min(max(max_length + 2, min_width), max_width)
            ws.column_dimensions[col_letter].width = adjusted_width

    def build_excel_with_collapsable_by_processo(self, df, sheet_name="Relatorio"):
        """
        Gera um Excel com agrupamento (collapsible) por PROCESSO,
        mantendo a PRIMEIRA linha visível em cada grupo.
        Retorna bytes prontos para attachment.
        """
        output = io.BytesIO()

        # 1️⃣ Escreve DF no Excel
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(
                writer,
                index=False,
                sheet_name=sheet_name
            )

        output.seek(0)

        # 2️⃣ Abre com openpyxl
        wb = load_workbook(output)
        ws = wb[sheet_name]

        # Descobrir índice da coluna "processo"
        headers = [cell.value for cell in ws[1]]
        processo_col = headers.index("processo") + 1  # openpyxl é 1-based

        current_processo = None
        first_row_of_group = None

        # 3️⃣ Agrupamento: esconder APENAS da segunda linha em diante
        for row in range(2, ws.max_row + 1):  # pula header
            processo = ws.cell(row=row, column=processo_col).value

            if processo != current_processo:
                # fecha grupo anterior
                if first_row_of_group and row - first_row_of_group > 1:
                    ws.row_dimensions.group(
                        first_row_of_group + 1,  # mantém a primeira visível
                        row - 1,
                        hidden=True
                    )

                current_processo = processo
                first_row_of_group = row

        # último grupo
        if first_row_of_group and ws.max_row - first_row_of_group >= 1:
            ws.row_dimensions.group(
                first_row_of_group + 1,
                ws.max_row,
                hidden=True
            )

        # 4️⃣ Configuração do outline
        ws.sheet_properties.outlinePr.summaryBelow = True
        self.auto_adjust_column_width(ws)
        self.apply_row_colors_by_qtd_alteracoes(ws)
        # 5️⃣ Salva Excel final
        final_output = io.BytesIO()
        wb.save(final_output)
        final_output.seek(0)

        return final_output.read()

    def should_send_report(self):
        result = self.queries.run_query(
            "SELECT status, last_sent FROM bronze.d_roll_monitor_expo_mailing LIMIT 1"
        )

        if not result:
            return False

        row = result[0]

        if row["status"] != "ACTIVE":
            return False

        now = datetime.now()
        last_sent = row["last_sent"]

        for run in config.RUNS:
            run_time = datetime.strptime(run, "%H:%M").time()
            slot_start = datetime.combine(now.date(), run_time)
            slot_end = slot_start + timedelta(minutes=config.WINDOW_MINUTES)

            # estamos dentro da janela?
            if slot_start <= now <= slot_end:
                # nunca foi enviado
                if last_sent is None:
                    return True

                # já enviou hoje nesse slot?
                if last_sent < slot_start:
                    return True

        return False


    def send(self):
        content = self.queries.get_diffs_from_db()
       
       # Manipula DF para o formato desejado
        df = pd.DataFrame(content)
        base = df[["processo", "campo", "armador","cliente","old_value", "new_value","email_responsavel", "email_resp_booking", "email_cs", "motivo_transferencia","booking","ds_quantidade_containers","porto_embarque"]]
        

        expanded = base.apply(self.expand_row, axis=1)
        expanded_df = pd.DataFrame(list(expanded))
            
        counts = (
            base.groupby("processo")
            .size()
            .rename("qtd_alteracoes")
        )

        final_df = expanded_df.merge(
            counts,
            on="processo",
            how="left"
        )

        final_df = final_df.sort_values(["processo"])
        final_df = final_df[
            [
                "processo",
                "armador",
                "cliente",
                "qtd_alteracoes",
                "campo",
                "navio_embarque (old)",
                "viagem (old)",
                "navio_embarque (new)",
                "viagem (new)",
                "previsao_embarque (old)",
                "previsao_embarque (new)",
                "navio_transbordo (old)",
                "navio_transbordo (new)",
                "previsao_embarque_transbordo (old)",
                "previsao_embarque_transbordo (new)",
                "booking",
                "ds_quantidade_containers",
                "porto_embarque",
                "email_responsavel",
                "email_resp_booking",
                "email_cs",
                "motivo_transferencia"
            ]
        ]
        # Manipula DF PARA EXCEL

        excel_bytes = self.build_excel_with_collapsable_by_processo(final_df)

        attachment = {
            "@odata.type": "#microsoft.graph.fileAttachment",
            "name": "relatorio_alteracoes.xlsx",
            "contentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "contentBytes": base64.b64encode(excel_bytes).decode("utf-8")
        }
        for recipient in self.recipient_email:
            self.mailing.send_mail(
                    to_address=recipient,
                    subject=self.subject,
                    html_content="Segue em anexo o relatório de alterações do Roll Monitor Exportação.",
                    attachments=[attachment])